import json
from openpyxl import Workbook

ECV = {
    "atmosphere": {
        "surface" : [
            "precipitation",
            "pressure",
            "radiation budget",
            "atmosphere temperature",
            "water vapour",
            "lightning",
            "upper-air temperature",
            "cloud"
        ],
        "atmospheric composition" : [
            "aerosol",
            "carbon dioxide",
            "methane",
            "ozone",
            "other greenhouse gases",
            "precursors for aerosols and ozone"
        ],
        "wind" : [
            "wind speed",
            "wind direction"
        ]
    },
    "land": {
        "hydrosphere": [
            "groundwater",
            "lake",
            "river discharge",
            "terrestrial water storage"
        ],
        "cryosphere": [
            "glacier",
            "ice sheet",
            "ice shelf",
            "permafrost",
            "snow"
        ],
        "biosphere": [
            "above-ground biomass",
            "albedo",
            "evaporation from land",
            "fire",
            "FAPAR",
            "land cover",
            "land surface temperature",
            "leaf area index",
            "soil carbon",
            "soil moisture"
        ],
        "anthroposphere": [
            "anthropogenic GHG fluxes",
            "anthropogenic water use"
        ]
    },
    "ocean": {
        "physical": [
            "ocean surface heat flux",
            "sea ice",
            "sea level",
            "sea state",
            "sea surface current",
            "sea surface salinity",
            "sea surface stress",
            "sea surface temperature",
            "subsurface current",
            "subsurface salinity",
            "subsurface temperature"
        ],
        "biogeochemical": [
            "inorganic carbon",
            "nitrous oxide",
            "nutrient",
            "ocean colour",
            "oxygen",
            "transient tracer"
        ],
        "biological/ecosystems": [
            "marine habitat",
            "plankton"
        ]
    }
}

ecv_aliases = {"precipitation" : [],
               "pressure" : [],
               "radiation budget" : [],
               "earth radiation budget" : [],
               "atmosphere temperature" : [],
               "wind speed" : [],
               "wind direction" : [],
               "lightning" : [],
               "upper-air temperature" : [],
               "water vapour" : ["water vapor"],
               "cloud" : [],
               "aerosol" : [],
            #    "carbon dioxide" : ["CO2", "CO 2", "CO₂"],
            #    "methane" : ["CH4", "CH 4", "CH₄"],
               "ozone" : ["O3", "O 3", "O₃"],
            #    "other greenhouse gases" : ["other greenhouse gas", "other GHG"],
               "precursors for aerosols and ozone" : ["precursor for aerosol", "precursors for aerosol", "precursor for ozone", "precursors for ozone"],
               "groundwater" : [],
               "lake" : [],
               "river discharge" : ["rivers discharge"],
               "terrestrial water storage" : [],
               "glacier" : [],
               "ice sheet" : [],
               "ice shelf" : ["ice shelves"],
               "permafrost" : [],
               "snow" : [],
               "above-ground biomass" : ["above ground biomass"],
               "albedo" : [],
               "evaporation from land" : ["evaporations from land"],
               "fire" : [],
               "FAPAR" : ["fraction of absorbed photosynthetically active radiation"],
               "land cover" : [],
               "land surface temperature" : ["lands surface temperature"],
               "leaf area index" : [],
               "soil carbon" : [],
               "soil moisture" : [],
               "anthropogenic GHG fluxes" : ["anthropogenic greenhouse gases fluxes", "anthropogenic greenhouse gas flux", "anthropogenic greenhouse gases flux", "anthropogenic greenhouse gas fluxes", "anthropogenic GHGs fluxes", "anthropogenic GHG flux", "anthropogenic GHGs flux"],
               "anthropogenic water use" : [],
               "ocean surface heat flux" : ["ocean surface heat fluxes"],
               "sea ice" : [],
               "sea level" : ["seas level"],
               "sea state" : ["seas state"],
               "sea surface current" : ["seas surface current", "seas surface currents"],
               "sea surface salinity" : ["seas surface salinity"],
               "sea surface stress" : ["sea surface stresses", "seas surface stress"],
               "sea surface temperature" : ["seas surface temperature"],
               "subsurface current" : [],
               "subsurface salinity" : ["subsurfaces salinity"],
               "subsurface temperature" : ["subsurfaces temperature"],
               "inorganic carbon" : [],
               "nitrous oxide" : ["N2O", "N 2O", "N₂O"],
               "nutrient" : [],
               "ocean colour" : ["oceans colour"],
               "oxygen" : [" O2", " O 2", " O₂"], # space to avoid 'o2' in 'co2'
               "transient tracer" : [],
               "marine habitat" : [],
               "plankton" : []
}

projects = ["Aerosol", "Biomass", "Climate Modelling User Group (CMUG)", "Cloud", "Fire", "Greenhouse Gases (GHGs)", "Glaciers", "High Resolution Land Cover", "Ice Sheets (Antarctic)", "Ice Sheets (Greenland)", "Lakes", "Land Cover", "Land Surface Temperature", "Ocean Colour", "Ozone", "Permafrost", "Precursors for aerosols and ozone", "RECCAP-2", "River Discharge", "Sea Ice", "Sea Level", "Sea Level Budget Closure", "Sea State", "Sea Surface Salinity", "Sea Surface Temperature", "Snow", "Soil Moisture", "Vegetation Parameters", "Water Vapour"]

project_aliases = {
               "aerosol" : [],
               "anthropogenic water use" : [],
               "biomass" : [],
               "CMUG" : ["Climate Modelling User Group", "climate modelling user group"],
               "cloud" : [],
               "fire" : [],
               "greenhouse gas" : [], # CO2 and CH4 are accounted for separately
               "glacier" : [],
               "high resolution land cover" : [],
               "ice sheet" : [],
               "lake" : [],
               "land cover" : [],
               "land surface temperature" : ["lands surface temperature"],
               "LOLIPOP" : ["LOng-LIved greenhouse gas PrOducts Performances", "long-lived greenhouse gas products performance"],
               "ocean colour" : ["oceans colour"],
               "ozone" : [" O3", " O 3", "O₃"],
               "permafrost" : [],
               "precursors for aerosols and ozone" : ["precursor for aerosol", "precursors for aerosol", "precursor for ozone", "precursors for ozone"],
               "RECCAP-2" : ['reccap', 'RECCAP'],
               "river discharge" : ["rivers discharge"],
               "sea ice" : [],
               "sea level" : ["seas level"],
               "sea level budget closure" : ["slbc", "sea level budget"],
               "sea state" : ["seas state"],
               "sea surface salinity" : ["seas surface salinity"],
               "sea surface temperature" : ["seas surface temperature"],
               "snow" : [],
               "soil moisture" : [],
               "vegetation parameters" : [],
               "water vapour" : ["water vapor"],
}

other_search_terms = {
    "carbon dioxide" : ["CO2", "CO 2", "CO₂"],
    "methane" : ["CH4", "CH 4", "CH₄"],
    "nitrous oxide" : ["N2O", "N 2O", "N₂O"],
    "other greenhouse gas" : ["other GHG"],
    "sulphur hexafluoride" : ["SF6", "SF 6", "SF₆"],
    "nitrogen dioxide" : ["NO2", "NO 2", "NO₂"],
    "sulphur dioxide" : ["SO2", "SO 2", "SO₂"],
    "carbon monoxide" : [],
    "formaldehyde" : ["HCHO"],
    "ammonia" : ["NH3", "NH 3", "NH₃"],
    "black carbon" : [],
    "particulate matter" : ['pm2.5', 'pm10'],
    "halogenated carbon compound" : ["chlorofluorocarbon", "hydrofluorocarbon", "hydrochlorofluorocarbon", "perfluorocarbon", "HFC", "PFC", "CFC"] # "HCFC" accounted for with "CFC"
}

context_search_terms = {
    "Copernicus Climate Change Service" : ['C3S'],
    "ERA-Interim" : ["ERA4", "ERA-4", "ERA5", "ERA-5"],
    "ECMWF" : ["European Centre for Medium-Range Weather Forecasts"],
    "EUMETSAT" : ["European Organisation for the Exploitation of Meteorological Satellites"]
}

cmip_search_terms = {
    'CFMIP': [],
    'DynVarMIP': [],
    'GMMIP': [],
    'HighResMIP': [],
    'PAMIP': [],
    #'OMIP': [],        # Check and edit extra occurences from GeoMIP, ScenarioMIP, PlioMIP,...
    'FAFMIP': [],
    'LS3MIP': [],
    'SIMIP' : [],
    'ISMIP6': [],
    'PMIP': [],
    'RFMIP': [],
    'DAMIP': [],
    'VolMIP': [],
    'AerChemMIP': [],
    'C4MIP': [],
    'LUMIP': [],
    'CDRMIP': [],
    'GeoMIP': [],
    'DCPP': [],
    'ScenarioMIP': [],
    'CORDEX': [],
    'VIACS AB': []
    }

sensors_search_terms = {
    'ATSR-2': [],
    'SMAP': [], 
    'MIRAS': [],
    'OLCI': [],
    'SAR-C': [],
    'AVHRR-3': [],
    'SSM/I': [],
    'AMSR2': [],
    'SMMR': [],
    'AMI-SCAT': [],
    'AMSRE': [],
    'ASCAT': [],
    'VIIRS': [],
    'WindSat': [],
    'AMI-SAR': [],
    'RA-2': [],
    'SeaWiFS': [],
    'AVHRR-2': [],
    'SLSTR': [],
    'Aquarius': [],
    'AltiKa': [],
    'SCIAMACHY': [],
    'SIRAL': [],
    'Poseidon-2': [],
    'Poseidon-3': [],
    'SSMI/S': [], 
    'TANSO-FTS': [],
    'VIRR': [],
    'ATSR': [],
    'SSALT': [],
    'PALSAR': [],
    'GRACE': [],
    'SRAL': [],
    'GFO-RA': [],
    'GOME-2': [],
    'PALSAR-2': [],
    'Poseidon-3B': [],
    'SAR-X': [],
    'TANSO-FTS-2': [],
    'AVHRR': [],
    'ETM+': [],
    'GOES-Imager': [],
    'GOMOS': [],
    'MWRI': [],
    'SEVIRI': [],
    'ACE-FTS': [], 
    'MIPAS': [],
    'OSIRIS': [],
    'SAR-2000': [],
    'TIRS': [],
    'TROPOMI': [],
    'MERIS': [],
    'ESMR': [],
    'IMAGER': [],
    'JAMI': [],
    'ASAR': [],
    'MSS': [],
    'MSI': [],
    'TMI': [],
    'TOMS': [],
    'OCO': [],
    'SMR': [],
    'OMI': [],
    'OLI': [],
    'ABI': [],
    'GMI': [],
    'GOME': [],
}

# Some sensors with potential miscount to check
# (e.g IMAGER can be spotted in 'imagery')
#'ESMR': [],
#'IMAGER': [],
#'JAMI': [],
#'ASAR': [],
#'MSS': [],
#'MSI': [],
#'TMI': [],
#'TOMS': [],
#'OCO': [],
#'SMR': [],
#'OMI': [],
#'OLI': [],
#'ABI': [],
#'GMI': [],
#'GOME': [],

satellite_search_terms = {
    'ERS-2': [],
    'Aqua': [],
    'ERS-1': [],
    'Sentinel-3A': [],
    'Metop-A': [],
    'SMOS': [],
    'Sentinel-3B': [],
    'Sentinel-1A': [],
    'Terra': [],
    'DMSP-F11': [],
    'Metop-B': [],
    'Nimbus-7': [],
    'Sentinel-1B': [],
    'DMSP-F08': [],
    'DMSP-F13': [],
    'GCOM-W1': [],
    'NOAA-16': [],
    'Coriolis': [],
    'Sentinel-2A': [],
    'TRMM': [],
    'CryoSat-2': [],
    'Orbview-2': [],
    'FY-3B': [],
    'NOAA-14': [],
    'NOAA-18': [],
    'NOAA-11': [],
    'NOAA-19': [],
    'NOAA-7': [],
    'NOAA-9': [],
    'SARAL': [],
    'SNPP': [],
    'Jason-1': [],
    'Jason-2': [],
    'NOAA-15': [],
    'NOAA-17': [],
    'SAC-D': [],
    'Sentinel-2B': [],
    'GOSAT': [],
    'GPM Core': [],
    'Jason-3': [],
    'NOAA-12': [],
    'GFO': [],
    'GOSAT-2': [],
    'GRACE': [],
    'Topex/Poseidon': [],
    'ALOS': [],
    'FY-3C': [],
    'FY-3D': [],
    'PROBA-V': [],
    'DMSP-F17': [],
    'ODIN': [],
    'ALOS-2': [],
    'Himawari-7': [],
    'Landsat-5': [],
    'Landsat-7': [],
    'Landsat-8': [],
    'SPOT-4': [],
    'Aura': [],
    'DMSP-F10': [],
    'DMSP-F14': [],
    'DMSP-F15': [],
    'DMSP-F16': [],
    'DMSP-F18': [],
    'GOES-12': [],
    'GOES-13': [],
    'GOES-16': [],
    'Himawari-6': [],
    'Landsat-4': [],
    'Meteosat-10': [],
    'Meteosat-11': [],
    'Meteosat-8': [],
    'Meteosat-9': [],
    'SPOT-5': [],
    'CSK-1': [],
    'CSK-2': [],
    'CSK-4': [],
    'Nimbus-5': [],
    'OCO-2': [],
    'RadarSat-2': [],
    'Sentinel-5P': []}

CMIP6 = {
    'Systematic Biases': {
        'Clouds/Circulation': ['CFMIP', 'DynVarMIP'],
        'Regional phenomena': ['GMMIP', 'HighResMIP', 'PAMIP'],
        'Ocean/Land/Ice': ['OMIP', 'FAFMIP', 'LS3MIP', 'SIMIP', 'ISMIP6']
    },
    'Response to Forcing': {
        'Paleo': ['PMIP'],
        'Characterizing forcing': ['RFMIP', 'DAMIP', 'VolMIP'],
        'Chemistry/Aerosols': ['AerChemMIP'],
        'Carbon cycle': ['C4MIP']
    },
    'Variability, Predictability, Future Scenarios': {
        'Land use': ['LUMIP'],
        'Geo-engineering': ['CDRMIP', 'GeoMIP'],
        'Decadal prediction': ['DCPP'],
        'Scenarios': ['ScenarioMIP'],
        'Impacts': ['CORDEX', 'VIACS AB']
    }
}

def merge_dicts(dict1, dict2):
    merged_dict = dict1.copy()  # Make a copy of the first dictionary
    merged_dict.update(dict2)    # Update with the second dictionary
    return merged_dict

search_terms = merge_dicts(ecv_aliases, project_aliases)
search_terms = merge_dicts(search_terms, other_search_terms)
search_terms = merge_dicts(search_terms, context_search_terms)
search_terms = merge_dicts(search_terms, cmip_search_terms)


with open("./data/cci/search_terms.json", "w") as f:
    json.dump(search_terms, f)

with open("./data/cci/ecv_aliases.json", "w") as f:
    json.dump(ecv_aliases, f)

with open("./data/cci/project_aliases.json", "w") as f:
    json.dump(project_aliases, f)

with open("./data/cci/ecv_classification.json", "w") as f:
    json.dump(ECV, f)

with open("./data/cci/projects.json", "w") as f:
    json.dump(projects, f)


def save_dict_to_excel(data_dict, file_name='output.xlsx'):
    # Create a new Excel workbook and grab the active worksheet
    wb = Workbook()
    ws = wb.active
    
    # Iterate over the dictionary items
    for row_index, (key, values) in enumerate(data_dict.items(), start=1):
        # Write the key in the first column
        ws.cell(row=row_index, column=1, value=key)
        
        # Write the values in the next columns
        for col_index, value in enumerate(values, start=2):
            ws.cell(row=row_index, column=col_index, value=value)
    
    # Save the workbook to the specified file name
    wb.save(file_name)
    print(f"Dictionary saved to '{file_name}' successfully.")

# Save the dictionary to an Excel file
save_dict_to_excel(search_terms, './temp/search_terms.xlsx')
