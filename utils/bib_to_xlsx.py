import bibtexparser
from openpyxl import Workbook

def convert_bib_to_xlsx(bib_file_path, xlsx_file_path):
    # Create a new workbook
    wb = Workbook()
    # Select the active worksheet
    ws = wb.active

    # Define column headers
    columns = ['Project', 'DOI', 'Title', 'Year', 'Author', 'Journal']
    # Write column headers to the first row
    for idx, col in enumerate(columns, start=1):
        ws.cell(row=1, column=idx, value=col)

    # Parse the BibTeX file
    with open(bib_file_path, 'r', encoding='utf-8') as bib_file:
        bib_database = bibtexparser.load(bib_file)

        # Iterate through entries
        for entry in bib_database.entries:
            # Initialize row data
            row_data = {}

            # Extract custom field 'project'
            project = entry.get('project', '')

            # Extract other fields with fallback for missing fields
            row_data['Project'] = project
            row_data['DOI'] = entry.get('doi', '')
            row_data['Title'] = entry.get('title', '')
            row_data['Year'] = entry.get('year', '')
            row_data['Author'] = entry.get('author', '')
            row_data['Journal'] = entry.get('journal', '')

            # Write row data to Excel sheet
            row_idx = len(ws['A']) + 1
            for col_idx, col_name in enumerate(columns, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ''))

    # Save the workbook to the specified path
    wb.save(xlsx_file_path)
    
# Example usage
convert_bib_to_xlsx('./data/cci/cci_papers_merged.bib', './data/cci/cci_papers_merged.xlsx')
